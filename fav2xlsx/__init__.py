import logging
import os, sys
from pathlib import Path
import json
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QLabel, QVBoxLayout, QTextEdit, QMessageBox, QFileDialog
from PyQt5.QtCore import Qt

# The openpyxl library allows for precise control of individual cells.
from openpyxl import Workbook


# logging setup
logger = logging.getLogger(__name__)
log_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'DEBUG.log')
# log_file_handler = logging.FileHandler(log_file_path, )
# noinspection SpellCheckingInspection
logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    datefmt='%m-%d %H:%M',
                    filename=log_file_path,
                    encoding='utf-8',  # Specify GB2312 encoding
                    filemode='w')
logger.info(f'Initializing logging - log file path: {log_file_path}\n')


class MyApp(QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Amap Favs (高德地图收藏夹位置) to Excel")  # Set window title
        self.setWindowState(Qt.WindowMaximized)  # Maximize the window

        self.textbox = QTextEdit("Paste the Amap favs to convert", self)
        self.button = QPushButton("Click Me", self)
        self.button.clicked.connect(self.submit_text)  # Connect the button click to a function
        self.label = QLabel("Input region", self)

        # Create a QVBoxLayout (vertical layout to organize widgets)
        layout = QVBoxLayout()
        layout.addWidget(self.textbox)
        layout.addWidget(self.button)
        layout.addWidget(self.label)

        self.setLayout(layout)  # Set the layout for the window

    def submit_text(self):
        # self.label.setText("Button Clicked!")  # Update the label text when the button is clicked
        user_input = self.textbox.toPlainText().strip()  # Get text from QTextEdit
        logger.info(f"Length of user input: {len(user_input)}")
        if user_input.strip():
            self.parse_amap(user_input)
        else:
            self.show_warning("Input cannot be empty!")
            return

    @staticmethod
    def show_warning(msg: str):
        # Create a message box
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)  # You can use different icons like Information, Warning, Critical, etc.
        msg.setText(msg)  # Message text
        msg.setWindowTitle("Warning")  # Window title
        msg.setStandardButtons(QMessageBox.Ok)  # Buttons to show (Ok)
        logger.warning(msg)

    def show_info(self, msg:str):
        # Show message in the bottom label
        self.label.setText(msg)

    def parse_amap(self, amap_fav: str):
        # check if input valid
        try:
            # Parse JSON string to Python dictionary
            fav_items = json.loads(amap_fav)["data"]["items"]
            del amap_fav
        except UnboundLocalError | json.JSONDecodeError:
            self.show_warning("Input not valid! Try again.")
            return
        # check if fav_items not empty
        if not (isinstance(fav_items, list) and fav_items):
            self.show_warning("No fav items found.")
        else:
            self.show_info(f"You have input {len(fav_items)} fav items.")

        # Create a new Workbook
        wb = Workbook()
        ws = wb.active  # Select the active worksheet
        # Write the headers to the first row (Row 1)
        headers = ['经度', '纬度', '名称', '地址', '收藏时间', '类别']
        for col_index, item in enumerate(headers, start=1):  # start=1 to start writing from column A
            ws.cell(row=1, column=col_index, value=item)  # Writing to row 1

        # iterate fav items
        for row_index, fav in enumerate(fav_items, start=2):
            try:
                ws.cell(row=row_index, column=headers.index('经度') + 1, value=fav["data"]["lon"])
                ws.cell(row=row_index, column=headers.index('纬度') + 1, value=fav["data"]["lat"])
                ws.cell(row=row_index, column=headers.index('名称') + 1, value=fav["data"]["name"])
                ws.cell(row=row_index, column=headers.index('地址') + 1, value=fav["data"]["address"])
                ws.cell(row=row_index, column=headers.index('收藏时间') + 1, value=fav["ts"])
            except Exception as e:
                # TODO handle data errors
                logger.warning(e)

        # Open file dialog
        excel_path, _ = QFileDialog.getSaveFileName(self, "Save File", "", "Excel files (*.xlsx);;All Files (*)")

        if excel_path:  # If the user selects a location
            # Save the workbook to the chosen path
            wb.save(excel_path)
            logger.info(f"Amap Favs saved to {excel_path}.")


if __name__ == '__main__':
    app = QApplication(sys.argv)  # Create the application
    window = MyApp()  # Create the main window
    window.show()  # Display the window
    sys.exit(app.exec_())  # Run the application's event loop


# TODO open this web link
